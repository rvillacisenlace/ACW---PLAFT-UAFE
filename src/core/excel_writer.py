"""
Capa de escritura del Excel. Dos implementaciones intercambiables:
- LocalExcelWriter: escribe directo al .xlsx local (para la demo, sin Graph API).
- GraphAPIWriter: cuando tengas tenant_id/client_id/client_secret/drive_id/item_id,
  se activa sin cambiar una línea de main.py.

main.py solo debe conocer la interfaz ExcelWriter, nunca la implementación.
"""
from abc import ABC, abstractmethod
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

class ExcelWriter(ABC):
    @abstractmethod
    def actualizar_estado_cliente(self, fila_excel: int, estado: str, detalle: str = "") -> None:
        raise NotImplementedError

    @abstractmethod
    def escribir_detalle_procesos(self, cliente, procesos_judiciales: list, denuncias: list) -> None:
        raise NotImplementedError

    @abstractmethod
    def escribir_procesos_omitidos(self, cliente, procesos_judiciales: list) -> None:
        raise NotImplementedError

    @abstractmethod
    def leer_clientes_pendientes(self, nombre_hoja: str = None) -> list:
        raise NotImplementedError

    @abstractmethod
    def leer_parametrizacion(self, nombre_hoja: str = "Parametrizacion") -> dict:
        raise NotImplementedError

    @abstractmethod
    def guardar(self) -> None:
        raise NotImplementedError


class LocalExcelWriter(ExcelWriter):
    """
    Apunta a la hoja real 'Revision' del Excel de trabajo. El encabezado
    real ocupa las filas 1-3 (grupo -> subgrupo -> campo, con celdas
    combinadas) - por eso el nombre de cada columna puede vivir en
    cualquiera de esas 3 filas, no solo en la fila 1. Los datos de
    clientes empiezan en la fila 4.
    """
    FILA_INICIO_DATOS = 4

    ESTILO_FUENTE = Font(name="Book Antiqua", size=11)
    ESTILO_ALINEACION = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ESTILO_RELLENO = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    COL_SRI_RAZON_SOCIAL = 14
    COL_SRI_ESTADO_CONTRIBUYENTE = 15
    COL_SRI_FECHA_INICIO = 16
    COL_SRI_FECHA_CESE = 17
    COL_SRI_FECHA_REINICIO = 18
    COL_SRI_CONTRIBUYENTE_FANTASMA = 19
    COL_SRI_TRANSACCIONES_INEXISTENTES = 20
    COL_SRI_ACTIVIDAD_ECONOMICA = 21

    COL_SRI_RL_RAZON_SOCIAL = 22
    COL_SRI_RL_ESTADO_CONTRIBUYENTE = 23
    COL_SRI_RL_FECHA_INICIO = 24
    COL_SRI_RL_FECHA_CESE = 25
    COL_SRI_RL_FECHA_REINICIO = 26
    COL_SRI_RL_ACTIVIDAD_ECONOMICA = 27
    COL_SRI_RL_CONTRIBUYENTE_FANTASMA = 28
    COL_SRI_RL_TRANSACCIONES_INEXISTENTES = 29

    COL_CONTRALORIA_DECLARACIONES = 132
    COL_CONTRALORIA_VIGENCIA = 133
    COL_CONTRALORIA_CARGO = 135
    COL_CONTRALORIA_TIEMPO = 136
    COL_CONTRALORIA_ULTIMO_ANIO = 137

    # Bloque SCVS Personas (Z-BK). "RUC", "Cargo", "Fecha de Constitucion",
    # "Observaciones" se repiten en muchas otras partes de la hoja -
    # indices directos, mismo criterio que los demas bloques.
    COL_SCVS_PERSONAS_TOTAL_PRESIDENTE_RL = 45  # AS
    COL_SCVS_PERSONAS_TOTAL_ACCIONISTA = 46  # AT
    COL_SCVS_PERSONAS_SLOTS = [
        (47, 48, 49, 50, 51, 52, 53, 54, 55),  # AU-BC
        (56, 57, 58, 59, 60, 61, 62, 63, 64),  # BD-BL
        (65, 66, 67, 68, 69, 70, 71, 72, 73),  # BM-BU
        (74, 75, 76, 77, 78, 79, 80, 81, 82),  # BV-CD
    ]

    COL_SENTENCIADOS_TOTAL = 143
    COL_SENTENCIADOS_SLOTS = [
        (144, 145, 146, 147),
        (148, 149, 150, 151),
        (152, 153, 154, 155),
    ]

    COL_CNJ_TOTAL = 156
    COL_CNJ_TEMATICA = 157
    COL_CNJ_SLOTS = [
        (158, 159, 160, 161, 162),
        (163, 164, 165, 166, 167),
        (168, 169, 170, 171, 172),
    ]

    def __init__(self, ruta_excel: str, nombre_hoja: str = "Revision"):
        self.ruta_excel = ruta_excel
        self.nombre_hoja = nombre_hoja
        self.wb = load_workbook(ruta_excel)
        self.hoja = self.wb[nombre_hoja]
        self.mapa_columnas = self._construir_mapa_columnas(self.hoja)

        import io
        self._imagenes_bytes_originales = []
        for hoja_wb in self.wb.worksheets:
            for img in getattr(hoja_wb, "_images", []):
                img.ref.seek(0)
                self._imagenes_bytes_originales.append((img, img.ref.read()))

    def _construir_mapa_columnas(self, hoja) -> dict:
        mapa = {}
        for col in range(1, hoja.max_column + 1):
            valor = (
                hoja.cell(row=3, column=col).value
                or hoja.cell(row=2, column=col).value
                or hoja.cell(row=1, column=col).value
            )
            if valor:
                mapa[str(valor).strip()] = col
        return mapa

    def _col(self, nombre_columna: str) -> int:
        if nombre_columna not in self.mapa_columnas:
            raise KeyError(f"Columna '{nombre_columna}' no encontrada en la hoja '{self.nombre_hoja}'")
        return self.mapa_columnas[nombre_columna]

    def _escribir_valor_con_estilo(self, fila: int, col: int, valor) -> None:
        celda = self.hoja.cell(row=fila, column=col)
        celda.value = str(valor).upper() if valor is not None else valor
        celda.font = self.ESTILO_FUENTE
        celda.alignment = self.ESTILO_ALINEACION
        celda.fill = self.ESTILO_RELLENO

    def leer_clientes_pendientes(self, nombre_hoja: str = None) -> list:
        from src.core.models import Cliente, TipoPersona

        col_id = self._col("Ruc / CI")
        col_nombres = self._col("Apellidos Y Nombres (P.Natural / P.Juridica)")
        col_razon = self._col("Razon Social (Empresa)")
        col_estado = self._col("ESTADO")

        clientes = []
        for fila in range(self.FILA_INICIO_DATOS, self.hoja.max_row + 1):
            identificacion = self.hoja.cell(row=fila, column=col_id).value
            if not identificacion or not str(identificacion).strip():
                continue

            estado = self.hoja.cell(row=fila, column=col_estado).value
            if estado and str(estado).strip().lower() == "completado":
                continue

            nombres = self.hoja.cell(row=fila, column=col_nombres).value
            razon = self.hoja.cell(row=fila, column=col_razon).value
            tipo = TipoPersona.NATURAL if nombres else TipoPersona.JURIDICA

            clientes.append(Cliente(
                identificacion=str(identificacion).strip(),
                tipo_persona=tipo,
                nombres_completos=(str(nombres).strip() if nombres else ""),
                razon_social=(str(razon).strip() if razon else ""),
                fila_excel=fila,
            ))

        return clientes

    def leer_parametrizacion(self, nombre_hoja: str = "Parametrizacion") -> dict:
        hoja = self.wb[nombre_hoja]
        parametros = {}
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            if len(fila) >= 2 and fila[0]:
                parametros[fila[0]] = fila[1]
        return parametros

    def escribir_antecedentes_penales(self, fila_excel: int, posee_antecedentes: bool) -> None:
        col = self._col("Posee antecedentes penales")
        self._escribir_valor_con_estilo(fila_excel, col, "SI" if posee_antecedentes else "NO")

    def escribir_sri_ruc(self, fila_excel: int, datos: dict, datos_representante_legal: dict = None) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RAZON_SOCIAL, datos.get("razon_social", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_ESTADO_CONTRIBUYENTE, datos.get("estado_contribuyente", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_INICIO, datos.get("fecha_inicio_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_CESE, datos.get("fecha_cese_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_REINICIO, datos.get("fecha_reinicio_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_CONTRIBUYENTE_FANTASMA, datos.get("contribuyente_fantasma", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_TRANSACCIONES_INEXISTENTES, datos.get("contribuyente_transacciones_inexistentes", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_ACTIVIDAD_ECONOMICA, datos.get("actividad_economica", "") or "-")

        self._escribir_valor_con_estilo(fila_excel, self._col("Representante Legal"), datos.get("representante_legal_nombre", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("ID Representante Legal"), datos.get("representante_legal_identificacion", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("Dirección Domicilio"), datos.get("direccion_matriz", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("Fecha de Constitucion"), datos.get("fecha_inicio_actividades", "") or "-")

        columnas_rl = [
            self.COL_SRI_RL_RAZON_SOCIAL, self.COL_SRI_RL_ESTADO_CONTRIBUYENTE,
            self.COL_SRI_RL_FECHA_INICIO, self.COL_SRI_RL_FECHA_CESE,
            self.COL_SRI_RL_FECHA_REINICIO, self.COL_SRI_RL_ACTIVIDAD_ECONOMICA,
            self.COL_SRI_RL_CONTRIBUYENTE_FANTASMA, self.COL_SRI_RL_TRANSACCIONES_INEXISTENTES,
        ]
        if not datos_representante_legal:
            for col in columnas_rl:
                self._escribir_valor_con_estilo(fila_excel, col, "-")
        else:
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_RAZON_SOCIAL, datos_representante_legal.get("razon_social", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_ESTADO_CONTRIBUYENTE, datos_representante_legal.get("estado_contribuyente", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_INICIO, datos_representante_legal.get("fecha_inicio_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_CESE, datos_representante_legal.get("fecha_cese_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_REINICIO, datos_representante_legal.get("fecha_reinicio_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_ACTIVIDAD_ECONOMICA, datos_representante_legal.get("actividad_economica", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_CONTRIBUYENTE_FANTASMA, datos_representante_legal.get("contribuyente_fantasma", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_TRANSACCIONES_INEXISTENTES, datos_representante_legal.get("contribuyente_transacciones_inexistentes", "") or "-")

    def escribir_sri_deudas(self, fila_excel: int, tiene_deuda_firme: bool, valor_deuda_firme: str = "") -> None:
        col = self._col("SRI DEUDAS")
        if tiene_deuda_firme:
            texto = f"El ciudadano / contribuyente registra un valor de deudas firmes de {valor_deuda_firme}"
        else:
            texto = "El ciudadano / contribuyente no registra deudas firmes."
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_sri_estado_tributario(self, fila_excel: int, resultado: str, obligaciones_pendientes: str = "") -> None:
        col = self._col("SRI OBLIGACIONES TRIBUTARIAS")
        if "AL DIA" in resultado.upper():
            texto = "AL DIA EN SUS OBLIGACIONES"
        else:
            texto = obligaciones_pendientes
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_municipios(self, fila_excel: int, resultados: dict) -> None:
        municipios_con_deuda = []
        total_deuda = 0.0
        algun_registrado = False

        for nombre_municipio, deuda in resultados.items():
            if deuda.registrado:
                algun_registrado = True
            if deuda.tiene_deuda:
                municipios_con_deuda.append(nombre_municipio)
            valor_limpio = (deuda.valor_total or "0").replace("$", "").replace(",", "").strip()
            try:
                total_deuda += float(valor_limpio)
            except ValueError:
                pass

        if not algun_registrado:
            texto_municipios = "No existen registros en los 5 municipios"
        elif municipios_con_deuda:
            texto_municipios = " / ".join(municipios_con_deuda)
        else:
            texto_municipios = "Ninguno"

        self._escribir_valor_con_estilo(fila_excel, self._col("MUNICIPIO"), texto_municipios)
        self._escribir_valor_con_estilo(fila_excel, self._col("DEUDA"), f"${total_deuda:,.2f}")

    def escribir_sercop_proveedor(self, fila_excel: int, estado: str) -> None:
        col = self._col("INCOP")
        self._escribir_valor_con_estilo(fila_excel, col, estado)

    def escribir_sercop_certificados(self, fila_excel: int, datos: dict) -> None:
        resultado_contratos = datos.get("contratos_pendientes", {}).get("resultado", "INDETERMINADO")
        resultado_incumplimientos = datos.get("incumplimientos", {}).get("resultado", "INDETERMINADO")
        self._escribir_valor_con_estilo(fila_excel, self._col("¿Tiene procesos pendientes con el Estado?"), resultado_contratos)
        self._escribir_valor_con_estilo(fila_excel, self._col("¿Es contratista incumplido?"), resultado_incumplimientos)

    def escribir_salud(self, fila_excel: int, situacion_laboral: str, tipo_afiliacion: str) -> None:
        self._escribir_valor_con_estilo(fila_excel, self._col("Situación Laboral"), situacion_laboral)
        self._escribir_valor_con_estilo(fila_excel, self._col("Tipo de Afiliación"), tipo_afiliacion)

    def escribir_iess(self, fila_excel: int, iess: str, deuda_obligaciones: str = "") -> None:
        col_iess = self._col("IESS")
        col_deuda = self._col("Deuda obligaciones patronales")
        self._escribir_valor_con_estilo(fila_excel, col_iess, iess)

        valor_limpio = (deuda_obligaciones or "").strip()
        if not valor_limpio:
            texto_deuda = "$0.00"
        else:
            texto_deuda = valor_limpio if valor_limpio.startswith("$") else f"${valor_limpio}"
        self._escribir_valor_con_estilo(fila_excel, col_deuda, texto_deuda)

    def escribir_scvs_companias(self, fila_excel: int, registrado: bool, cumplimiento_obligaciones: str = "") -> None:
        col = self._col("Supercias")
        if not registrado:
            texto = "-"
        elif cumplimiento_obligaciones.strip().upper().startswith("SI"):
            texto = "SI Cumple con sus obligaciones"
        else:
            texto = "NO Cumple con sus obligaciones"
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_contraloria(self, fila_excel: int, resumen: dict) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_DECLARACIONES, resumen.get("posee_declaraciones", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_VIGENCIA, resumen.get("vigencia", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_CARGO, resumen.get("cargo", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_TIEMPO, resumen.get("tiempo", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_ULTIMO_ANIO, resumen.get("ultimo_anio_en_cargo", "-"))

    def escribir_scvs_personas(self, fila_excel: int, resultado: dict, nombre_persona_relacionada: str) -> None:
        """
        resultado: dict devuelto por ScraperSCVSPersonas.buscar_cliente()
        ({"total_presidente_rl", "total_accionista", "participaciones"}).
        nombre_persona_relacionada: nombre de quien se busco (cliente o
        representante legal) - se repite igual en cada slot ocupado,
        confirmado con Excel real de referencia.
        """
        self._escribir_valor_con_estilo(fila_excel, self.COL_SCVS_PERSONAS_TOTAL_PRESIDENTE_RL, resultado.get("total_presidente_rl", 0))
        self._escribir_valor_con_estilo(fila_excel, self.COL_SCVS_PERSONAS_TOTAL_ACCIONISTA, resultado.get("total_accionista", 0))

        participaciones = resultado.get("participaciones", [])
        for i, columnas_slot in enumerate(self.COL_SCVS_PERSONAS_SLOTS):
            col_id, col_empresa, col_ruc, col_cargo, col_capital, col_situacion, col_fecha, col_obs, col_patrimonio = columnas_slot
            if i < len(participaciones):
                p = participaciones[i]
                self._escribir_valor_con_estilo(fila_excel, col_id, nombre_persona_relacionada)
                self._escribir_valor_con_estilo(fila_excel, col_empresa, p.nombre_empresa)
                self._escribir_valor_con_estilo(fila_excel, col_ruc, p.ruc_empresa)
                self._escribir_valor_con_estilo(fila_excel, col_cargo, p.cargo)
                self._escribir_valor_con_estilo(fila_excel, col_capital, p.capital_invertido)
                self._escribir_valor_con_estilo(fila_excel, col_situacion, p.situacion_legal)
                self._escribir_valor_con_estilo(fila_excel, col_fecha, p.fecha_constitucion)
                self._escribir_valor_con_estilo(fila_excel, col_obs, p.actividad_economica)
                self._escribir_valor_con_estilo(fila_excel, col_patrimonio, p.patrimonio_ultimo_anio)
            else:
                for col in columnas_slot:
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_sentenciados(self, fila_excel: int, total_encontrado: int, top3: list) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_SENTENCIADOS_TOTAL, str(total_encontrado))
        for i, (col_no, col_proceso, col_fecha, col_infraccion) in enumerate(self.COL_SENTENCIADOS_SLOTS):
            if i < len(top3):
                s = top3[i]
                self._escribir_valor_con_estilo(fila_excel, col_no, str(i + 1))
                self._escribir_valor_con_estilo(fila_excel, col_proceso, s.numero_proceso)
                self._escribir_valor_con_estilo(fila_excel, col_fecha, s.fecha_resolucion)
                self._escribir_valor_con_estilo(fila_excel, col_infraccion, s.infraccion)
            else:
                for col in (col_no, col_proceso, col_fecha, col_infraccion):
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_funcion_judicial(self, fila_excel: int, procesos: list, total_procesos: int, tematica_general: str) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_CNJ_TOTAL, str(total_procesos))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CNJ_TEMATICA, tematica_general)

        procesos_detalle = [p for p in procesos if not p.omitido_por_volumen and not p.excluido_por_materia]

        for i, (col_no, col_fecha, col_proceso, col_accion, col_obs) in enumerate(self.COL_CNJ_SLOTS):
            if i < len(procesos_detalle):
                p = procesos_detalle[i]
                self._escribir_valor_con_estilo(fila_excel, col_no, str(i + 1))
                self._escribir_valor_con_estilo(fila_excel, col_fecha, p.fecha_ingreso)
                self._escribir_valor_con_estilo(fila_excel, col_proceso, p.numero_proceso)
                self._escribir_valor_con_estilo(fila_excel, col_accion, p.accion_infraccion_delito)
                self._escribir_valor_con_estilo(fila_excel, col_obs, p.resumen_ia)
            else:
                for col in (col_no, col_fecha, col_proceso, col_accion, col_obs):
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_contraloria_resumen_general(self, fila_excel: int, resumen_general: str) -> None:
        col = self._col("CONTRALORÍA")
        self._escribir_valor_con_estilo(fila_excel, col, resumen_general)

    def escribir_fiscalia_resumen_general(self, fila_excel: int, resumen_general: str) -> None:
        col = self._col("FISCALIA")
        self._escribir_valor_con_estilo(fila_excel, col, resumen_general)

    def escribir_estado_final(self, fila_excel: int, resultados: dict) -> None:
        requiere_revision = any(
            isinstance(r, dict) and r.get("requiere_revision_manual")
            for r in resultados.values()
        )
        estado = "Completado con pendientes" if requiere_revision else "Completado"
        self.actualizar_estado_cliente(fila_excel, estado)

    def actualizar_estado_cliente(self, fila_excel: int, estado: str, detalle: str = "") -> None:
        col = self._col("ESTADO")
        self._escribir_valor_con_estilo(fila_excel, col, estado)

    def escribir_sitios_a_revisar(self, fila_excel: int, texto: str) -> None:
        col = self._col("SITIOS A REVISAR")
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_ruta_evidencia(self, fila_excel: int, ruta: str) -> None:
        col = self._col("RUTA EVIDENCIA")
        self._escribir_valor_con_estilo(fila_excel, col, ruta)

    def escribir_detalle_procesos(self, cliente, procesos_judiciales: list, denuncias: list) -> None:
        raise NotImplementedError(
            "escribir_detalle_procesos no esta implementado en LocalExcelWriter "
            "(depende de metodos que solo existen en GraphAPIWriter) - pendiente."
        )

    def escribir_procesos_omitidos(self, cliente, procesos_judiciales: list) -> None:
        raise NotImplementedError(
            "escribir_procesos_omitidos no esta implementado en LocalExcelWriter "
            "(depende de metodos que solo existen en GraphAPIWriter) - pendiente."
        )

    def _refrescar_streams_de_imagenes(self) -> None:
        import io
        for img, datos in self._imagenes_bytes_originales:
            img.ref = io.BytesIO(datos)

    def guardar(self) -> None:
        self._refrescar_streams_de_imagenes()
        self.wb.save(self.ruta_excel)


import truststore
truststore.inject_into_ssl()

import os
import msal
import requests


class GraphAPIWriter(ExcelWriter):
    """
    Escribe directamente en el Excel real alojado en el OneDrive de la
    cuenta de servicio, vía Microsoft Graph API. Usa sesión de workbook
    para agrupar los cambios y evitar bloquear el archivo a otros usuarios.
    """

    def __init__(self, cuenta_onedrive: str, drive_id: str, item_id: str,
                 tenant_id: str, client_id: str, client_secret: str,
                 nombre_hoja: str = "Revision"):
        self.drive_id = drive_id
        self.item_id = item_id
        self.nombre_hoja = nombre_hoja

        self._app_msal = msal.ConfidentialClientApplication(
            client_id=client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            client_credential=client_secret,
        )
        self._refrescar_token()
        self.base_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook"
        self.session_id = self._crear_sesion()
        self._mapa_columnas = self._leer_encabezados()

    def _crear_sesion(self) -> str:
        resp = requests.post(
            f"{self.base_url}/createSession",
            headers=self.headers,
            json={"persistChanges": True},
        )
        resp.raise_for_status()
        return resp.json()["id"]

    def _headers_con_sesion(self) -> dict:
        self._refrescar_token()
        return {**self.headers, "workbook-session-id": self.session_id}

    def _patch_con_reintento_sesion(self, url: str, json_body: dict):
        """
        PATCH con recuperación automática de sesión expirada. Confirmado
        con evidencia real (2026-09-02): en corridas largas (18 sitios
        antes de llegar a escribir el Excel), la sesión de Graph API
        puede expirar/invalidarse por un error transitorio del lado del
        servidor ("InvalidSession"). Se recrea la sesión UNA vez y se
        reintenta antes de fallar del todo.
        """
        resp = requests.patch(url, headers=self._headers_con_sesion(), json=json_body)
        if resp.status_code == 400 and "InvalidSession" in resp.text:
            print("    [Graph API] Sesión expirada - recreando sesión y reintentando...")
            self.session_id = self._crear_sesion()
            resp = requests.patch(url, headers=self._headers_con_sesion(), json=json_body)
        return resp

    def _leer_encabezados(self) -> dict:
        url = f"{self.base_url}/worksheets/{self.nombre_hoja}/range(address='1:3')"
        resp = requests.get(url, headers=self._headers_con_sesion())
        resp.raise_for_status()
        valores = resp.json()["values"]

        fila1 = valores[0] if len(valores) > 0 else []
        fila2 = valores[1] if len(valores) > 1 else []
        fila3 = valores[2] if len(valores) > 2 else []
        num_columnas = max(len(fila1), len(fila2), len(fila3))

        mapa = {}
        for idx in range(num_columnas):
            valor = None
            for fila in (fila3, fila2, fila1):
                if idx < len(fila) and fila[idx]:
                    valor = fila[idx]
                    break
            if valor:
                mapa[str(valor).strip()] = idx
        return mapa

    def _col(self, nombre_columna: str) -> int:
        if nombre_columna not in self._mapa_columnas:
            raise KeyError(f"Columna '{nombre_columna}' no encontrada en la hoja '{self.nombre_hoja}'")
        return self._mapa_columnas[nombre_columna]

    def _asegurar_hoja_existe(self, nombre_hoja: str, encabezados: list) -> None:
        resp = requests.get(f"{self.base_url}/worksheets", headers=self._headers_con_sesion())
        resp.raise_for_status()
        nombres_existentes = [h["name"] for h in resp.json()["value"]]

        if nombre_hoja not in nombres_existentes:
            requests.post(
                f"{self.base_url}/worksheets/add",
                headers=self._headers_con_sesion(),
                json={"name": nombre_hoja},
            ).raise_for_status()

            letra_final = chr(ord("A") + len(encabezados) - 1)
            requests.patch(
                f"{self.base_url}/worksheets/{nombre_hoja}/range(address='A1:{letra_final}1')",
                headers=self._headers_con_sesion(),
                json={"values": [encabezados]},
            ).raise_for_status()

    def _obtener_ultima_fila(self, nombre_hoja: str) -> int:
        resp = requests.get(
            f"{self.base_url}/worksheets/{nombre_hoja}/usedRange",
            headers=self._headers_con_sesion(),
        )
        resp.raise_for_status()
        datos = resp.json()
        return datos["rowIndex"] + datos["rowCount"]

    def _escribir_fila(self, nombre_hoja: str, num_fila: int, valores: list) -> None:
        letra_final = chr(ord("A") + len(valores) - 1)
        rango = f"A{num_fila}:{letra_final}{num_fila}"
        resp = requests.patch(
            f"{self.base_url}/worksheets/{nombre_hoja}/range(address='{rango}')",
            headers=self._headers_con_sesion(),
            json={"values": [valores]},
        )
        resp.raise_for_status()

    def _escribir_valor_con_estilo(self, fila_excel: int, col_idx_0based: int, valor) -> None:
        """
        Escribe el valor de una celda Y le aplica el estilo completo
        (alineacion, fuente, relleno) en cada llamada. Cada peticion
        pasa por _patch_con_reintento_sesion, que recupera sola la
        sesion si expiro a mitad de una corrida larga (confirmado con
        evidencia real: "InvalidSession" tras ~15 min de scraping antes
        de llegar a escribir el Excel).
        """
        letra_columna = self._indice_a_letra(col_idx_0based)
        celda = f"{letra_columna}{fila_excel}"
        texto = str(valor).upper() if valor is not None else ""
        url_rango = f"{self.base_url}/worksheets/{self.nombre_hoja}/range(address='{celda}')"

        resp = self._patch_con_reintento_sesion(url_rango, {
            "values": [[texto]],
            "horizontalAlignment": "Center", "verticalAlignment": "Center", "wrapText": True,
        })
        resp.raise_for_status()

        self._patch_con_reintento_sesion(f"{url_rango}/format/font", {"name": "Book Antiqua", "size": 11}).raise_for_status()
        self._patch_con_reintento_sesion(f"{url_rango}/format/fill", {"color": "#D9D9D9"}).raise_for_status()

    @staticmethod
    def _indice_a_letra(indice_0based: int) -> str:
        indice = indice_0based
        letra = ""
        while True:
            indice, resto = divmod(indice, 26)
            letra = chr(65 + resto) + letra
            if indice == 0:
                break
            indice -= 1
        return letra

    def actualizar_estado_cliente(self, fila_excel: int, estado: str, detalle: str = "") -> None:
        col_idx = self._col("ESTADO")
        self._escribir_valor_con_estilo(fila_excel, col_idx, estado)

    def escribir_antecedentes_penales(self, fila_excel: int, posee_antecedentes: bool) -> None:
        col = self._col("Posee antecedentes penales")
        self._escribir_valor_con_estilo(fila_excel, col, "SI" if posee_antecedentes else "NO")

    def escribir_sitios_a_revisar(self, fila_excel: int, texto: str) -> None:
        col = self._col("SITIOS A REVISAR")
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_ruta_evidencia(self, fila_excel: int, ruta: str) -> None:
        col = self._col("RUTA EVIDENCIA")
        self._escribir_valor_con_estilo(fila_excel, col, ruta)

    COL_SRI_RAZON_SOCIAL = 13
    COL_SRI_ESTADO_CONTRIBUYENTE = 14
    COL_SRI_FECHA_INICIO = 15
    COL_SRI_FECHA_CESE = 16
    COL_SRI_FECHA_REINICIO = 17
    COL_SRI_CONTRIBUYENTE_FANTASMA = 18
    COL_SRI_TRANSACCIONES_INEXISTENTES = 19
    COL_SRI_ACTIVIDAD_ECONOMICA = 20

    COL_SRI_RL_RAZON_SOCIAL = 21
    COL_SRI_RL_ESTADO_CONTRIBUYENTE = 22
    COL_SRI_RL_FECHA_INICIO = 23
    COL_SRI_RL_FECHA_CESE = 24
    COL_SRI_RL_FECHA_REINICIO = 25
    COL_SRI_RL_ACTIVIDAD_ECONOMICA = 26
    COL_SRI_RL_CONTRIBUYENTE_FANTASMA = 27
    COL_SRI_RL_TRANSACCIONES_INEXISTENTES = 28

    COL_CONTRALORIA_DECLARACIONES = 131
    COL_CONTRALORIA_VIGENCIA = 132
    COL_CONTRALORIA_CARGO = 134
    COL_CONTRALORIA_TIEMPO = 135
    COL_CONTRALORIA_ULTIMO_ANIO = 136

    COL_SCVS_PERSONAS_TOTAL_PRESIDENTE_RL = 44
    COL_SCVS_PERSONAS_TOTAL_ACCIONISTA = 45
    COL_SCVS_PERSONAS_SLOTS = [
        (46, 47, 48, 49, 50, 51, 52, 53, 54),
        (55, 56, 57, 58, 59, 60, 61, 62, 63),
        (64, 65, 66, 67, 68, 69, 70, 71, 72),
        (73, 74, 75, 76, 77, 78, 79, 80, 81),
    ]

    COL_SENTENCIADOS_TOTAL = 142
    COL_SENTENCIADOS_SLOTS = [
        (143, 144, 145, 146),
        (147, 148, 149, 150),
        (151, 152, 153, 154),
    ]

    COL_CNJ_TOTAL = 155
    COL_CNJ_TEMATICA = 156
    COL_CNJ_SLOTS = [
        (157, 158, 159, 160, 161),
        (162, 163, 164, 165, 166),
        (167, 168, 169, 170, 171),
    ]

    def escribir_sri_ruc(self, fila_excel: int, datos: dict, datos_representante_legal: dict = None) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RAZON_SOCIAL, datos.get("razon_social", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_ESTADO_CONTRIBUYENTE, datos.get("estado_contribuyente", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_INICIO, datos.get("fecha_inicio_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_CESE, datos.get("fecha_cese_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_FECHA_REINICIO, datos.get("fecha_reinicio_actividades", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_CONTRIBUYENTE_FANTASMA, datos.get("contribuyente_fantasma", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_TRANSACCIONES_INEXISTENTES, datos.get("contribuyente_transacciones_inexistentes", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_ACTIVIDAD_ECONOMICA, datos.get("actividad_economica", "") or "-")

        self._escribir_valor_con_estilo(fila_excel, self._col("Representante Legal"), datos.get("representante_legal_nombre", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("ID Representante Legal"), datos.get("representante_legal_identificacion", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("Dirección Domicilio"), datos.get("direccion_matriz", "") or "-")
        self._escribir_valor_con_estilo(fila_excel, self._col("Fecha de Constitucion"), datos.get("fecha_inicio_actividades", "") or "-")

        columnas_rl = [
            self.COL_SRI_RL_RAZON_SOCIAL, self.COL_SRI_RL_ESTADO_CONTRIBUYENTE,
            self.COL_SRI_RL_FECHA_INICIO, self.COL_SRI_RL_FECHA_CESE,
            self.COL_SRI_RL_FECHA_REINICIO, self.COL_SRI_RL_ACTIVIDAD_ECONOMICA,
            self.COL_SRI_RL_CONTRIBUYENTE_FANTASMA, self.COL_SRI_RL_TRANSACCIONES_INEXISTENTES,
        ]
        if not datos_representante_legal:
            for col in columnas_rl:
                self._escribir_valor_con_estilo(fila_excel, col, "-")
        else:
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_RAZON_SOCIAL, datos_representante_legal.get("razon_social", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_ESTADO_CONTRIBUYENTE, datos_representante_legal.get("estado_contribuyente", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_INICIO, datos_representante_legal.get("fecha_inicio_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_CESE, datos_representante_legal.get("fecha_cese_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_FECHA_REINICIO, datos_representante_legal.get("fecha_reinicio_actividades", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_ACTIVIDAD_ECONOMICA, datos_representante_legal.get("actividad_economica", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_CONTRIBUYENTE_FANTASMA, datos_representante_legal.get("contribuyente_fantasma", "") or "-")
            self._escribir_valor_con_estilo(fila_excel, self.COL_SRI_RL_TRANSACCIONES_INEXISTENTES, datos_representante_legal.get("contribuyente_transacciones_inexistentes", "") or "-")

    def escribir_sri_deudas(self, fila_excel: int, tiene_deuda_firme: bool, valor_deuda_firme: str = "") -> None:
        col = self._col("SRI DEUDAS")
        if tiene_deuda_firme:
            texto = f"El ciudadano / contribuyente registra un valor de deudas firmes de {valor_deuda_firme}"
        else:
            texto = "El ciudadano / contribuyente no registra deudas firmes."
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_sri_estado_tributario(self, fila_excel: int, resultado: str, obligaciones_pendientes: str = "") -> None:
        col = self._col("SRI OBLIGACIONES TRIBUTARIAS")
        if "AL DIA" in resultado.upper():
            texto = "AL DIA EN SUS OBLIGACIONES"
        else:
            texto = obligaciones_pendientes
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_municipios(self, fila_excel: int, resultados: dict) -> None:
        municipios_con_deuda = []
        total_deuda = 0.0
        algun_registrado = False

        for nombre_municipio, deuda in resultados.items():
            if deuda.registrado:
                algun_registrado = True
            if deuda.tiene_deuda:
                municipios_con_deuda.append(nombre_municipio)
            valor_limpio = (deuda.valor_total or "0").replace("$", "").replace(",", "").strip()
            try:
                total_deuda += float(valor_limpio)
            except ValueError:
                pass

        if not algun_registrado:
            texto_municipios = "No existen registros en los 5 municipios"
        elif municipios_con_deuda:
            texto_municipios = " / ".join(municipios_con_deuda)
        else:
            texto_municipios = "Ninguno"

        self._escribir_valor_con_estilo(fila_excel, self._col("MUNICIPIO"), texto_municipios)
        self._escribir_valor_con_estilo(fila_excel, self._col("DEUDA"), f"${total_deuda:,.2f}")

    def escribir_sercop_proveedor(self, fila_excel: int, estado: str) -> None:
        col = self._col("INCOP")
        self._escribir_valor_con_estilo(fila_excel, col, estado)

    def escribir_sercop_certificados(self, fila_excel: int, datos: dict) -> None:
        resultado_contratos = datos.get("contratos_pendientes", {}).get("resultado", "INDETERMINADO")
        resultado_incumplimientos = datos.get("incumplimientos", {}).get("resultado", "INDETERMINADO")
        self._escribir_valor_con_estilo(fila_excel, self._col("¿Tiene procesos pendientes con el Estado?"), resultado_contratos)
        self._escribir_valor_con_estilo(fila_excel, self._col("¿Es contratista incumplido?"), resultado_incumplimientos)

    def escribir_salud(self, fila_excel: int, situacion_laboral: str, tipo_afiliacion: str) -> None:
        self._escribir_valor_con_estilo(fila_excel, self._col("Situación Laboral"), situacion_laboral)
        self._escribir_valor_con_estilo(fila_excel, self._col("Tipo de Afiliación"), tipo_afiliacion)

    def escribir_iess(self, fila_excel: int, iess: str, deuda_obligaciones: str = "") -> None:
        col_iess = self._col("IESS")
        col_deuda = self._col("Deuda obligaciones patronales")
        self._escribir_valor_con_estilo(fila_excel, col_iess, iess)

        valor_limpio = (deuda_obligaciones or "").strip()
        if not valor_limpio:
            texto_deuda = "$0.00"
        else:
            texto_deuda = valor_limpio if valor_limpio.startswith("$") else f"${valor_limpio}"
        self._escribir_valor_con_estilo(fila_excel, col_deuda, texto_deuda)

    def escribir_scvs_companias(self, fila_excel: int, registrado: bool, cumplimiento_obligaciones: str = "") -> None:
        col = self._col("Supercias")
        if not registrado:
            texto = "-"
        elif cumplimiento_obligaciones.strip().upper().startswith("SI"):
            texto = "SI Cumple con sus obligaciones"
        else:
            texto = "NO Cumple con sus obligaciones"
        self._escribir_valor_con_estilo(fila_excel, col, texto)

    def escribir_contraloria(self, fila_excel: int, resumen: dict) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_DECLARACIONES, resumen.get("posee_declaraciones", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_VIGENCIA, resumen.get("vigencia", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_CARGO, resumen.get("cargo", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_TIEMPO, resumen.get("tiempo", "-"))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CONTRALORIA_ULTIMO_ANIO, resumen.get("ultimo_anio_en_cargo", "-"))

    def escribir_scvs_personas(self, fila_excel: int, resultado: dict, nombre_persona_relacionada: str) -> None:
        """
        resultado: dict devuelto por ScraperSCVSPersonas.buscar_cliente()
        ({"total_presidente_rl", "total_accionista", "participaciones"}).
        nombre_persona_relacionada: nombre de quien se busco (cliente o
        representante legal) - se repite igual en cada slot ocupado,
        confirmado con Excel real de referencia.
        """
        self._escribir_valor_con_estilo(fila_excel, self.COL_SCVS_PERSONAS_TOTAL_PRESIDENTE_RL, resultado.get("total_presidente_rl", 0))
        self._escribir_valor_con_estilo(fila_excel, self.COL_SCVS_PERSONAS_TOTAL_ACCIONISTA, resultado.get("total_accionista", 0))

        participaciones = resultado.get("participaciones", [])
        for i, columnas_slot in enumerate(self.COL_SCVS_PERSONAS_SLOTS):
            col_id, col_empresa, col_ruc, col_cargo, col_capital, col_situacion, col_fecha, col_obs, col_patrimonio = columnas_slot
            if i < len(participaciones):
                p = participaciones[i]
                self._escribir_valor_con_estilo(fila_excel, col_id, nombre_persona_relacionada)
                self._escribir_valor_con_estilo(fila_excel, col_empresa, p.nombre_empresa)
                self._escribir_valor_con_estilo(fila_excel, col_ruc, p.ruc_empresa)
                self._escribir_valor_con_estilo(fila_excel, col_cargo, p.cargo)
                self._escribir_valor_con_estilo(fila_excel, col_capital, p.capital_invertido)
                self._escribir_valor_con_estilo(fila_excel, col_situacion, p.situacion_legal)
                self._escribir_valor_con_estilo(fila_excel, col_fecha, p.fecha_constitucion)
                self._escribir_valor_con_estilo(fila_excel, col_obs, p.actividad_economica)
                self._escribir_valor_con_estilo(fila_excel, col_patrimonio, p.patrimonio_ultimo_anio)
            else:
                for col in columnas_slot:
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_sentenciados(self, fila_excel: int, total_encontrado: int, top3: list) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_SENTENCIADOS_TOTAL, str(total_encontrado))
        for i, (col_no, col_proceso, col_fecha, col_infraccion) in enumerate(self.COL_SENTENCIADOS_SLOTS):
            if i < len(top3):
                s = top3[i]
                self._escribir_valor_con_estilo(fila_excel, col_no, str(i + 1))
                self._escribir_valor_con_estilo(fila_excel, col_proceso, s.numero_proceso)
                self._escribir_valor_con_estilo(fila_excel, col_fecha, s.fecha_resolucion)
                self._escribir_valor_con_estilo(fila_excel, col_infraccion, s.infraccion)
            else:
                for col in (col_no, col_proceso, col_fecha, col_infraccion):
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_funcion_judicial(self, fila_excel: int, procesos: list, total_procesos: int, tematica_general: str) -> None:
        self._escribir_valor_con_estilo(fila_excel, self.COL_CNJ_TOTAL, str(total_procesos))
        self._escribir_valor_con_estilo(fila_excel, self.COL_CNJ_TEMATICA, tematica_general)

        procesos_detalle = [p for p in procesos if not p.omitido_por_volumen and not p.excluido_por_materia]

        for i, (col_no, col_fecha, col_proceso, col_accion, col_obs) in enumerate(self.COL_CNJ_SLOTS):
            if i < len(procesos_detalle):
                p = procesos_detalle[i]
                self._escribir_valor_con_estilo(fila_excel, col_no, str(i + 1))
                self._escribir_valor_con_estilo(fila_excel, col_fecha, p.fecha_ingreso)
                self._escribir_valor_con_estilo(fila_excel, col_proceso, p.numero_proceso)
                self._escribir_valor_con_estilo(fila_excel, col_accion, p.accion_infraccion_delito)
                self._escribir_valor_con_estilo(fila_excel, col_obs, p.resumen_ia)
            else:
                for col in (col_no, col_fecha, col_proceso, col_accion, col_obs):
                    self._escribir_valor_con_estilo(fila_excel, col, "-")

    def escribir_contraloria_resumen_general(self, fila_excel: int, resumen_general: str) -> None:
        col = self._col("CONTRALORÍA")
        self._escribir_valor_con_estilo(fila_excel, col, resumen_general)

    def escribir_fiscalia_resumen_general(self, fila_excel: int, resumen_general: str) -> None:
        col = self._col("FISCALIA")
        self._escribir_valor_con_estilo(fila_excel, col, resumen_general)

    def escribir_estado_final(self, fila_excel: int, resultados: dict) -> None:
        requiere_revision = any(
            isinstance(r, dict) and r.get("requiere_revision_manual")
            for r in resultados.values()
        )
        estado = "Completado con pendientes" if requiere_revision else "Completado"
        self.actualizar_estado_cliente(fila_excel, estado)

    def escribir_detalle_procesos(self, cliente, procesos_judiciales: list, denuncias: list) -> None:
        encabezados = [
            "Identificacion_Cliente", "Sitio", "Numero_Proceso",
            "Tipo_Proceso", "Demandado_o_Rol", "Lugar", "Resumen_IA",
        ]
        self._asegurar_hoja_existe("DetalleProcesos", encabezados)

        procesos_filtrados = [
            p for p in procesos_judiciales
            if not p.omitido_por_volumen and not p.excluido_por_materia
        ]
        denuncias_filtradas = [d for d in denuncias if d.nombre_sospechoso]

        fila_actual = self._obtener_ultima_fila("DetalleProcesos") + 1

        if not procesos_filtrados:
            self._escribir_fila("DetalleProcesos", fila_actual, [
                cliente.identificacion, "Sitio 1 - Función Judicial",
                "", "", "", "", "Sin procesos vigentes",
            ])
            fila_actual += 1
        else:
            for proceso in procesos_filtrados:
                self._escribir_fila("DetalleProcesos", fila_actual, [
                    cliente.identificacion, "Sitio 1 - Función Judicial",
                    proceso.numero_proceso,
                    proceso.materia or proceso.accion_infraccion_delito,
                    proceso.demandado, proceso.lugar,
                    proceso.resumen_ia or "(resumen no generado)",
                ])
                fila_actual += 1

        for denuncia in denuncias_filtradas:
            self._escribir_fila("DetalleProcesos", fila_actual, [
                cliente.identificacion, "Sitio 2 - Fiscalía",
                denuncia.numero_noticia_delito, denuncia.delito,
                denuncia.nombre_sospechoso, denuncia.lugar, "",
            ])
            fila_actual += 1

    def escribir_procesos_omitidos(self, cliente, procesos_judiciales: list) -> None:
        encabezados = [
            "Identificacion_Cliente", "Numero_Proceso",
            "Accion_Infraccion", "Demandado", "Estado_Observacion",
        ]
        self._asegurar_hoja_existe("ProcesosOmitidos", encabezados)

        omitidos = [
            p for p in procesos_judiciales
            if p.omitido_por_volumen or p.excluido_por_materia
        ]
        if not omitidos:
            return

        fila_inicial = self._obtener_ultima_fila("ProcesosOmitidos") + 1
        filas_valores = [
            [cliente.identificacion, p.numero_proceso, p.accion_infraccion_delito,
             p.demandado, p.resumen_ia]
            for p in omitidos
        ]
        self._escribir_filas_batch("ProcesosOmitidos", fila_inicial, filas_valores)

    def leer_clientes_pendientes(self, nombre_hoja: str = None) -> list:
        from src.core.models import Cliente, TipoPersona

        hoja = nombre_hoja or self.nombre_hoja
        resp = requests.get(
            f"{self.base_url}/worksheets/{hoja}/usedRange",
            headers=self._headers_con_sesion(),
        )
        resp.raise_for_status()
        valores = resp.json()["values"]

        idx_id = self._col("Ruc / CI")
        idx_nombres = self._col("Apellidos Y Nombres (P.Natural / P.Juridica)")
        idx_razon = self._col("Razon Social (Empresa)")
        idx_estado = self._col("ESTADO")

        clientes = []
        for offset, fila in enumerate(valores[3:], start=4):
            identificacion_cruda = fila[idx_id] if idx_id < len(fila) else None
            if not identificacion_cruda or not str(identificacion_cruda).strip():
                continue

            estado = fila[idx_estado] if idx_estado < len(fila) else ""
            if estado and str(estado).strip().lower() == "completado":
                continue

            nombres = fila[idx_nombres] if idx_nombres < len(fila) else ""
            razon = fila[idx_razon] if idx_razon < len(fila) else ""
            tipo = TipoPersona.NATURAL if nombres else TipoPersona.JURIDICA

            clientes.append(Cliente(
                identificacion=str(identificacion_cruda).strip(),
                tipo_persona=tipo,
                nombres_completos=(str(nombres).strip() if nombres else ""),
                razon_social=(str(razon).strip() if razon else ""),
                fila_excel=offset,
            ))

        return clientes

    def leer_parametrizacion(self, nombre_hoja: str = "Parametrizacion") -> dict:
        resp = requests.get(
            f"{self.base_url}/worksheets/{nombre_hoja}/usedRange",
            headers=self._headers_con_sesion(),
        )
        resp.raise_for_status()
        valores = resp.json()["values"]

        parametros = {}
        for fila in valores[1:]:
            if len(fila) >= 2 and fila[0]:
                parametros[fila[0]] = fila[1]
        return parametros

    def _escribir_filas_batch(self, nombre_hoja: str, fila_inicial: int, filas_valores: list) -> None:
        if not filas_valores:
            return

        num_columnas = len(filas_valores[0])
        letra_final = chr(ord("A") + num_columnas - 1)
        fila_final = fila_inicial + len(filas_valores) - 1
        rango = f"A{fila_inicial}:{letra_final}{fila_final}"

        resp = requests.patch(
            f"{self.base_url}/worksheets/{nombre_hoja}/range(address='{rango}')",
            headers=self._headers_con_sesion(),
            json={"values": filas_valores},
        )
        resp.raise_for_status()

    def guardar(self) -> None:
        try:
            requests.post(f"{self.base_url}/closeSession", headers=self._headers_con_sesion())
        except Exception:
            pass

    def _refrescar_token(self) -> None:
        resultado = self._app_msal.acquire_token_for_client(
            scopes=["https://graph.microsoft.com/.default"]
        )
        if "access_token" not in resultado:
            raise RuntimeError(f"No se pudo renovar el token de Graph API: {resultado.get('error_description')}")

        self.headers = {"Authorization": f"Bearer {resultado['access_token']}"}